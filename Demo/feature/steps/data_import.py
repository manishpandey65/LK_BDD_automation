from openpyxl import *

book = load_workbook ( '../feature/Data/login.xlsx' )

ws = book.get_active_sheet ( )
rowcount = 2
# global data_login_email
# gobal data_login_password
for r in ws.rows:
    status = ws.cell ( row=rowcount , column=4 )  # storing individual cell data
    if (status.value == "true"):  # user wants to login in Daphne so store the ops username and pssword
        user_name = ws.cell ( row=rowcount , column=2 ).value
        pass_word = ws.cell ( row=rowcount , column=3 ).value
        #data_login_email = user_name
        #data_login_password = pass_word.value
        print(user_name)
    #print("no")
    rowcount = rowcount + 1

"""
def data_read():
    ws = book.get_active_sheet ( )
    rowcount=1
    colcount=1
    read=[]
    for r in ws.rows:
        for c in ws.columns:
            read=ws.cell(row=rowcount,column=colcount)
        print(read)
    print("sucessfully done")
    return read



def data_import(app):
    book = load_workbook ( 'E:/projects/lk_Demo/features/Data/login.xlsx' )
    ws = book.get_active_sheet ( )
    rowcount = 1
    #global data_login_email
    #gobal data_login_password
    for r in ws.rows:
        application = ws.cell ( row=rowcount , column=1 )  # storing individual cell data
        if (application.value == app):  # user wants to login in Daphne so store the ops username and pssword
            user_name = ws.cell ( row=rowcount , column=2 )
            pass_word = ws.cell ( row=rowcount , column=3 )
            data_login_email = user_name.value
            data_login_password = pass_word.value
            #print(data_login_email)
        rowcount = rowcount + 1
    return data_login_email,data_login_password


"""
